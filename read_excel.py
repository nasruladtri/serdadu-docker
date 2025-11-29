import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\rendy\Desktop\serdadu-docker\database\dataset\dkb_s2_2024.xlsx')

result = {
    'sheet_names': wb.sheetnames,
    'sheets': {}
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = []
    
    # Get first row (headers)
    for cell in ws[1]:
        headers.append(cell.value)
    
    result['sheets'][sheet_name] = {
        'headers': headers
    }

print(json.dumps(result, indent=2, ensure_ascii=False))
